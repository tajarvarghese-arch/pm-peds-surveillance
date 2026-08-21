#!/usr/bin/env python3
"""Generate a SYNTHETIC master workbook for testing the browser loader.

The real workbook holds confidential PM Pediatrics operating data and is never
committed. This script writes a file with the same sheet names, headers and
row shapes, filled with invented numbers, so `js/volumes.js::parseMasterWorkbook`
can be exercised end to end without a real export present.

    python scripts/make_synthetic_master.py [out.xlsx]

Default output is private/synthetic_master.xlsx, which is gitignored.
"""
from __future__ import annotations

import math
import random
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

RNG = random.Random(20260821)          # deterministic: same fixture every run
BANNER = Font(name="Arial", size=9, italic=True, color="555555")
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F4E79")
BODY = Font(name="Arial", size=10)

CATEGORIES = [
    ("Seasonal", "Respiratory conditions", True),
    ("Seasonal", "Pharyngitis & Strep pharyngitis", True),
    ("Seasonal", "Fever", True),
    ("Seasonal", "Otitis media", True),
    ("Seasonal", "Viral infections", True),
    ("Non-seasonal", "Skin conditions", False),
    ("Non-seasonal", "Allergic reactions", False),
    ("Non-seasonal", "Genitourinary system conditions", False),
    ("Injury", "Injury (all)", False),
    ("Uncategorized", "Uncategorized", False),
]
ICDS = [
    ("J02.0", "Streptococcal pharyngitis", "Strep pharyngitis"),
    ("H66.001", "Acute suppurative otitis media, right ear", "Otitis media"),
    ("J05.0", "Acute obstructive laryngitis [croup]", "Croup"),
    ("J10.1", "Influenza with other respiratory manifestations", "Influenza"),
    ("J18.9", "Pneumonia, unspecified organism", "Pneumonia"),
    ("N39.0", "Urinary tract infection, site not specified", "UTI"),
]
REGIONS = ["Region Alpha", "Region Bravo", "Region Charlie", "Region Delta"]
SITES = [f"Site {chr(65 + i)}{i:02d}" for i in range(24)]
MARKETS = ["northeast", "corridor", "expansion"]


def mondays(start: date, weeks: int) -> list[date]:
    start -= timedelta(days=start.weekday())
    return [start + timedelta(weeks=i) for i in range(weeks)]


def season(d: date, amp: float = 0.55) -> float:
    """1.0 at the summer trough, ~1+2*amp at the winter peak."""
    doy = d.timetuple().tm_yday
    return 1.0 + amp * (1 + math.cos(2 * math.pi * (doy - 15) / 365.25))


def sheet(wb: Workbook, name: str, note: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=f"SYNTHETIC TEST DATA — {note}").font = BANNER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(3, len(headers)))
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font, c.fill = HEAD, FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[ws.cell(row=2, column=j).column_letter].width = max(12, len(h) + 3)
    for i, row in enumerate(rows, start=3):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BODY
            if isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A3"
    return ws


def build(out: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    weeks = mondays(date(2025, 1, 6), 80)
    months = [date(y, m, 1) for y in (2025, 2026) for m in range(1, 13)][:20]

    # 01 — weekly diagnoses
    dx_rows, weekly_total, weekly_inf = [], {}, {}
    for w in weeks:
        s = season(w)
        for grp, cat, infectious in CATEGORIES:
            base = {"Seasonal": 4200, "Non-seasonal": 900, "Injury": 3000, "Uncategorized": 420}[grp]
            v = int(base * (s if infectious else 1.0) * RNG.uniform(0.9, 1.1))
            dx_rows.append([w, grp, cat, v, infectious])
            weekly_total[w] = weekly_total.get(w, 0) + v
            if infectious:
                weekly_inf[w] = weekly_inf.get(w, 0) + v
    sheet(wb, "01_Weekly_Diagnoses", "week x diagnosis category",
          ["week", "seasonality_group", "diagnosis_category", "visits", "infectious"], dx_rows)

    # 02 — weekly high-acuity
    ha_rows, weekly_ha = [], {}
    for w in weeks:
        s = season(w)
        for code, desc, grp in ICDS:
            v = int(900 * s * RNG.uniform(0.8, 1.2))
            ha_rows.append([w, code, desc, grp, v])
            weekly_ha[w] = weekly_ha.get(w, 0) + v
    sheet(wb, "02_Weekly_HighAcuity", "week x ICD-10 code",
          ["week", "icd_code", "icd_description", "clinical_grouping", "diagnoses"], ha_rows)

    # 03 — weekly channel
    ch_rows, chan = [], {}
    for w in weeks:
        tot = weekly_total[w]
        walk = int(tot * RNG.uniform(0.45, 0.53))
        pre = tot - walk
        pphr = round(tot / RNG.uniform(6800, 7600), 2)
        newp = int(tot * RNG.uniform(0.24, 0.28))
        hours = round(tot / pphr)
        ch_rows.append([w, walk, pre, pphr, newp, tot, hours,
                        round(walk / tot * 100, 2), round(newp / tot * 100, 2)])
        chan[w] = (walk, pre, pphr, newp)
    sheet(wb, "03_Weekly_Channel", "week",
          ["week", "walkin_visits", "prebooked_visits", "patients_per_operating_hour",
           "new_patients", "total_visits", "implied_operating_hours",
           "walkin_share_pct", "new_patient_share_pct"], ch_rows)

    # 04 — weekly funnel by region (percentages stored as fractions)
    fu_rows = []
    for w in weeks:
        for r in REGIONS:
            slots = int(2600 * RNG.uniform(0.9, 1.1))
            rate = round(RNG.uniform(0.38, 0.70), 3)
            pre = int(slots * rate * RNG.uniform(0.95, 1.05))
            fu_rows.append([r, w, slots, rate, round(1 - pre / slots, 4), pre,
                            round(RNG.uniform(0.02, 0.05), 4), round(RNG.uniform(0.09, 0.14), 4),
                            round(RNG.uniform(0.001, 0.008), 4), round(RNG.uniform(0.01, 0.02), 4),
                            RNG.randint(0, 14), RNG.randint(2, 28), RNG.randint(1, 20)])
    sheet(wb, "04_Weekly_Funnel_Regional", "week x region",
          ["region", "week", "available_slots", "booking_rate", "pct_potentially_unused_slots",
           "prebooked_visits", "pct_no_show", "pct_cancelled", "pct_improperly_cancelled",
           "pct_left_without_being_seen", "cancel_went_to_other_pm",
           "cancel_service_not_offered", "cancel_nonpar_or_inactive_insurance"], fu_rows)

    # 05 — monthly locations (wide)
    hdr = ["Location"] + [m.strftime("%b %Y") for m in months] + ["Total"]
    loc_rows = []
    for s_ in SITES:
        vals = [int(900 * season(m) * RNG.uniform(0.85, 1.15)) for m in months]
        loc_rows.append([s_] + vals + [sum(vals)])
    sheet(wb, "05_Monthly_Locations", "site x month (wide)", hdr, loc_rows)

    # 06 — monthly new patients
    np_rows = []
    for m in months:
        cur = int(30000 * season(m) * RNG.uniform(0.9, 1.1))
        prev = int(cur * RNG.uniform(1.05, 1.25)) if m.year == 2026 else None
        np_rows.append([m, cur, prev,
                        (cur - prev) if prev else None,
                        round((cur / prev - 1) * 100, 2) if prev else None])
    sheet(wb, "06_Monthly_NewPatients", "month",
          ["month", "new_patients_by_patient_id", "prior_year_same_month",
           "yoy_change", "yoy_pct"], np_rows)

    # 07 — weekly site visits
    sw_rows = []
    for w in weeks:
        for i, s_ in enumerate(SITES):
            sw_rows.append([w, s_, "2021 & Prior" if i % 3 else "2022 & After",
                            int(230 * season(w) * RNG.uniform(0.8, 1.2))])
    sheet(wb, "07_Weekly_Site_Visits", "week x site",
          ["week", "site", "growth_assumption", "visits"], sw_rows)

    # 08 — daily BH / telehealth
    bh_rows = []
    d0 = date(2025, 1, 1)
    for i in range((date(2026, 1, 18) - d0).days + 1):
        d = d0 + timedelta(days=i)
        bh_rows.append([d, int(3600 * season(d) * RNG.uniform(0.8, 1.2)),
                        int(120 * season(d, 0.9) * RNG.uniform(0.7, 1.3)),
                        int(60 * (1 + i / 400) * RNG.uniform(0.6, 1.4))])
    sheet(wb, "08_Daily_BH_Telehealth", "day",
          ["date", "urgent_care_visits", "telemedicine_visits", "behavioral_health_visits"], bh_rows)

    # 09 — site master
    sm_rows = []
    for i, s_ in enumerate(SITES):
        yoy = round(RNG.uniform(-20, 14), 1)
        sm_rows.append([s_, ["NY", "NJ", "MD", "FL", "TX"][i % 5], MARKETS[i % 3],
                        round(RNG.uniform(1.5, 15), 1), RNG.randint(4000, 15000), yoy,
                        yoy > 0, round(38 + RNG.uniform(0, 4), 3), round(-78 + RNG.uniform(0, 6), 3)])
    sheet(wb, "09_Site_Master", "site",
          ["site", "state", "market_group", "tenure_years_at_jan_2026", "visits_jan_jul_2025",
           "yoy_pct_jan_jul", "grew_2026", "lat", "lon"], sm_rows)

    # 10 — reference totals
    sheet(wb, "10_Reference_Totals", "metric (no time dimension)",
          ["Metric", "Value", "Window", "Source"],
          [["Urgent Care Visits", sum(weekly_total.values()), "synthetic", "synthetic"],
           ["Walk-In Visits", sum(c[0] for c in chan.values()), "synthetic", "synthetic"],
           ["Pre-Booked Visits", sum(c[1] for c in chan.values()), "synthetic", "synthetic"]])

    # 11 — derived weekly
    der_rows = []
    for w in weeks:
        tot = weekly_total[w]
        inf = weekly_inf[w]
        walk, pre, pphr, newp = chan[w]
        ha = weekly_ha[w]
        der_rows.append([w, int(tot * 0.62), int(tot * 0.16), int(tot * 0.18), int(tot * 0.04),
                         tot, inf, tot - inf, ha, round(ha / tot * 1000, 1),
                         round(tot * 0.62 / tot * 100, 2), walk, pre, newp, pphr])
    sheet(wb, "11_Derived_Weekly", "week (80 complete weeks)",
          ["week", "seasonal_visits", "non_seasonal_visits", "injury_visits", "uncategorized_visits",
           "total_visits", "infectious_visits", "non_infectious_visits", "high_acuity_diagnoses",
           "high_acuity_per_1000_visits", "pct_seasonal", "walkin_visits", "prebooked_visits",
           "new_patients", "patients_per_operating_hour"], der_rows)

    # 12 — site metadata
    sheet(wb, "12_Site_Metadata", "site",
          ["site", "workday_location", "open_date", "cohort", "growth_assumption"],
          [[s_, f"XX {s_}", date(2012 + i % 12, 1 + i % 12, 15),
            f"{2012 + i % 12} Openings", "2021 & Prior" if i % 3 else "2022 & After"]
           for i, s_ in enumerate(SITES)])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(wb.sheetnames)} sheets)")
    print("SYNTHETIC DATA — invented numbers, safe to share. Never replace with a real export.")


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("private/synthetic_master.xlsx")
    build(target)
